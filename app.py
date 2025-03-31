
import re
from datetime import datetime
from io import BytesIO
import pandas as pd
import zipfile
import streamlit as st
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Page setup
st.set_page_config(page_title="Coast2coast Roster Generator", page_icon="📋")
st.title("Daily Class Roster Generator")

uploaded_file = st.file_uploader("Upload Bookeo Export (.xlsx)", type=["xlsx"])
sample_file = "Mar 27_Markham_Blended_Recert_Patrick.xlsx"

if uploaded_file:
    df = pd.read_excel(uploaded_file, sheet_name=0)

    # Identify required columns
    start_time_col = "Start"
    course_type_col = df.columns[12]
    course_level_col = "Courses & Levels"
    textbook_col = "Textbook"
    kit_col = "First Aid Kits"

    def extract_location_and_category(text):
        match = re.match(r"(.+?)\s*\((.+)\)", str(text))
        if match:
            category = match.group(1).strip()
            location = match.group(2).strip()
            if "First Aid & CPR/AED" in category:
                category = "First Aid & CPR/AED"
            return category, location
        return "Unknown", "Unknown"

    df[['Course Category', 'Location']] = df[course_type_col].apply(lambda x: pd.Series(extract_location_and_category(x)))
    df['Start Time'] = pd.to_datetime(df[start_time_col])
    grouped = df.groupby(['Location', 'Course Category', 'Start Time'])

    # Load and trim the info block
    info_block = pd.read_excel(sample_file, sheet_name="Roster", header=None).iloc[15:36]
    trimmed_block = info_block.iloc[8:-2].reset_index(drop=True)

    # Style presets
    header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    orange_fill = PatternFill(start_color="FDBE87", end_color="FDBE87", fill_type="solid")
    blue_bar = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for (location, category, start_time), group in grouped:
            wb = Workbook()
            ws = wb.active
            ws.title = "Roster"

            # Build roster data
            roster_data = pd.DataFrame({
                "Pass (P) or Fail (F) or No-show (N)": "",
                "First Name": group["First name (participant)"],
                "Last Name": group["Last name (participant)"],
                "Course Taken": group[course_level_col],
                "QR Code / Paper Test\n(specify which)": "",
                "First Aid Kit\n(HIGHLIGHT)\ndelivered = green\nnot delivered = red": group[kit_col].fillna("").apply(lambda x: "" if "No Thanks! I Will Risk It Without A First Aid Kit." in str(x) else str(x)),
                "Textbook\n(HIGHLIGHT)\ndelivered = green\nnot delivered = red": group[textbook_col].apply(lambda x: "Purchased" if pd.notna(x) and str(x).strip() != "" else ""),
                "Notes for Instructor": ""
            })

            # Add students to worksheet
            for r_idx, row in enumerate(dataframe_to_rows(roster_data, index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.alignment = center_align
                    cell.border = border
                    if r_idx == 1:
                        cell.fill = header_fill
                        cell.font = Font(bold=True)
                    elif r_idx % 2 == 0:
                        cell.fill = alt_fill

            # NOTE TO OFFICE in A17
            note_cell = ws.cell(row=17, column=1, value="NOTE TO OFFICE: >>>")
            note_cell.fill = orange_fill
            note_cell.font = Font(bold=True)
            note_cell.alignment = center_align
            note_cell.border = border

            # Blank orange block B16:H19
            ws.merge_cells(start_row=16, start_column=2, end_row=19, end_column=8)
            for row in ws.iter_rows(min_row=16, max_row=19, min_col=2, max_col=8):
                for cell in row:
                    cell.fill = orange_fill
                    cell.border = border

            # Instruction rows 21–23
            instructions = [
                "Additional Instructions to Follow:",
                "Recertification students MUST show a copy of their current (valid) certification. Standard First Aid must be from the Canadian Red Cross. CPR/AED Level C can be from any WSIB-approved provider.",
                "Inform the office staff of any changes (course upgrades, spelling of names, why an item was not delivered to a participant, etc.)"
            ]
            for i, text in enumerate(instructions):
                row_num = 21 + i
                ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
                cell = ws.cell(row=row_num, column=1, value=text)
                cell.fill = blue_bar
                cell.font = Font(bold=True)
                cell.alignment = center_align
                cell.border = border

            # Final info block starting row 25
            for i, row in trimmed_block.iterrows():
                for j, value in enumerate(row):
                    cell = ws.cell(row=25 + i, column=j + 1, value=value)
                    cell.alignment = center_align
                    cell.border = border
                    cell.fill = blue_bar
                    cell.font = Font(bold=True)

            # Name and save the file
            date_str = start_time.strftime("%b %d")
            time_str = start_time.strftime("%I%p").lstrip('0')
            clean_course = re.sub(r'\W+', '', category.replace(" ", ""))
            clean_location = re.sub(r'\W+', '', location)
            file_name = f"{date_str}_{clean_location}_{clean_course}_{time_str}.xlsx"
            temp_buffer = BytesIO()
            wb.save(temp_buffer)
            zipf.writestr(file_name, temp_buffer.getvalue())

    st.success("Rosters generated!")
    st.download_button("📥 Download All Rosters (ZIP)", data=zip_buffer.getvalue(),
                       file_name="Rosters.zip", mime="application/zip")
